"""
KassandraOpenAI Sistem Testi v2.1
=================================
Tüm test sorularını /admin/test-chat endpoint'ine gönderir ve detaylı rapor üretir.
HEM TÜRKÇE HEM İNGİLİZCE soruları destekler.
GELİŞTİRİLMİŞ SKORLAMA SİSTEMİ - Daha adil ve akıllı değerlendirme.

Kullanım:
    python test_runner.py              # Sadece TR soruları test et
    python test_runner.py --both       # Hem TR hem EN soruları test et
    python test_runner.py --en         # Sadece EN soruları test et

Çıktılar:
    - test_results_{timestamp}.json (ham veri)
    - test_report_{timestamp}.xlsx (filtrelenebilir tablo)
    - test_report_{timestamp}.md (okunabilir döküman - TAM YANITLAR)
"""

import json
import httpx
import asyncio
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any
import time
import sys
import re

# Konfigürasyon
BASE_URL = "http://localhost:8000"
TEST_ENDPOINT = "/admin/test-chat"
QUESTIONS_FILE = Path(__file__).parent / "test_questions.json"
OUTPUT_DIR = Path(__file__).parent / "reports"

# ============================================================
# GELİŞTİRİLMİŞ SKORLAMA SİSTEMİ v2.1
# ============================================================

# Türkçe anahtar kelimeler
KEYWORDS_TR = {
    # Pozitif - yardımcı ve olumlu yanıtlar (+5 puan)
    "positive": [
        "evet", "tabii", "tabii ki", "mümkün", "yapabiliriz", "sağlayabiliriz",
        "edebilirsiniz", "yapabilirsiniz", "alabilirsiniz", "bulunmaktadır",
        "mevcuttur", "var", "sunuyoruz", "hizmet veriyoruz"
    ],
    
    # Selamlama - profesyonel giriş (+5 puan)
    "greeting": [
        "merhaba", "hoş geldiniz", "günaydın", "iyi günler", "iyi akşamlar"
    ],
    
    # Kapanış - profesyonel sonlandırma (+5 puan)
    "closing": [
        "iyi günler", "iyi akşamlar", "görüşmek üzere", "rica ederim",
        "memnuniyetle", "yardımcı olmaktan mutluluk", "başka bir soru",
        "yardımcı olabilir miyim", "iyi tatiller", "iyi yolculuklar"
    ],
    
    # Kibar red - profesyonel ama olumsuz (CEZA YOK, nötr)
    "polite_decline": [
        "maalesef", "üzgünüm", "ne yazık ki", "mümkün değil",
        "bulunmamaktadır", "hizmet vermemekteyiz", "kabul etmiyoruz"
    ],
    
    # Yönlendirme - müşteri temsilcisine aktarma (+3 puan)
    "redirect": [
        "canlı müşteri temsilcisi", "temsilcimize", "yönetim", "resepsiyon",
        "bizi arayın", "arayabilirsiniz", "ulaşabilirsiniz", "iletişime geçin"
    ],
    
    # Bilgilendirici içerik - somut bilgi (+10 puan)
    "informative_patterns": [
        r"\d{1,2}:\d{2}",           # Saat formatı (08:00, 14:30)
        r"\d{1,2}\s*(nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım)",  # Tarih
        r"\d+\s*€",                  # Euro fiyat
        r"\d+\s*TL",                 # TL fiyat
        r"\d+\s*(metre|km|dakika|saat|gece|kişi)",  # Ölçü birimleri
        r"(tek yön|çift yön|dahil|hariç)",  # Fiyat detayları
    ],
    
    # Gerçek hata mesajları - sistem hatası (-20 puan)
    "error": [
        "hata oluştu", "bir sorun oluştu", "tekrar deneyin", "error",
        "bağlantı hatası", "sistem hatası", "işlem yapılamadı",
        "anlayamadım", "anlamadım", "tekrar yazar mısınız"
    ],
    
    # Profesyonel ifadeler (+3 puan)
    "professional": [
        "yardımcı olabilirim", "bilgi vermek isterim", "detay paylaşayım",
        "şöyle açıklayayım", "kısaca belirteyim"
    ]
}

# İngilizce anahtar kelimeler
KEYWORDS_EN = {
    "positive": [
        "yes", "of course", "certainly", "available", "we can", "we offer",
        "you can", "we have", "we provide", "included"
    ],
    
    "greeting": [
        "hello", "hi", "welcome", "good morning", "good afternoon", "good evening"
    ],
    
    "closing": [
        "have a nice day", "best regards", "kind regards", "happy to help",
        "let me know", "feel free to", "don't hesitate"
    ],
    
    "polite_decline": [
        "unfortunately", "sorry", "we're sorry", "regret", "not available",
        "we don't", "we cannot", "isn't possible"
    ],
    
    "redirect": [
        "contact us", "call us", "reach out", "customer service",
        "reception", "front desk"
    ],
    
    "informative_patterns": [
        r"\d{1,2}:\d{2}",           # Time format
        r"\d{1,2}\s*(april|may|june|july|august|september|october|november)",
        r"€\s*\d+",                  # Euro price
        r"\$\s*\d+",                 # Dollar price
        r"\d+\s*(meters|km|minutes|hours|nights|people|guests)",
    ],
    
    "error": [
        "error occurred", "something went wrong", "try again", "error",
        "connection failed", "system error", "couldn't process",
        "didn't understand", "please repeat"
    ],
    
    "professional": [
        "happy to help", "let me explain", "i'd like to inform",
        "please note", "for your information"
    ]
}


def evaluate_response_v21(question: str, response: str, lang: str = "tr") -> Dict:
    """
    GELİŞTİRİLMİŞ SKORLAMA SİSTEMİ v2.1
    
    Başlangıç: 50 puan
    Maksimum: 100 puan
    Minimum: 0 puan
    """
    response_lower = response.lower()
    question_lower = question.lower()
    flags = []
    score = 50  # Başlangıç skoru
    details = []  # Skor detayları
    
    # Dil bazlı keyword seçimi
    keywords = KEYWORDS_TR if lang == "tr" else KEYWORDS_EN
    
    # ============================================================
    # 1. UZUNLUK DEĞERLENDİRMESİ
    # ============================================================
    response_len = len(response)
    
    if response_len < 20:
        score -= 25
        flags.append("TOO_SHORT")
        details.append("Çok kısa yanıt: -25")
    elif response_len < 50:
        score -= 10
        flags.append("SHORT")
        details.append("Kısa yanıt: -10")
    elif 50 <= response_len <= 500:
        score += 15
        flags.append("GOOD_LENGTH")
        details.append("Uygun uzunluk: +15")
    elif response_len > 500:
        score += 10
        flags.append("DETAILED")
        details.append("Detaylı yanıt: +10")
    
    # ============================================================
    # 2. SELAMLAMA KONTROLÜ
    # ============================================================
    has_greeting = any(g in response_lower for g in keywords["greeting"])
    if has_greeting:
        score += 5
        flags.append("HAS_GREETING")
        details.append("Selamlama var: +5")
    
    # ============================================================
    # 3. KAPANIŞ KONTROLÜ
    # ============================================================
    has_closing = any(c in response_lower for c in keywords["closing"])
    if has_closing:
        score += 5
        flags.append("HAS_CLOSING")
        details.append("Profesyonel kapanış: +5")
    
    # ============================================================
    # 4. POZİTİF İÇERİK
    # ============================================================
    positive_count = sum(1 for p in keywords["positive"] if p in response_lower)
    if positive_count > 0:
        bonus = min(positive_count * 3, 10)  # Max +10
        score += bonus
        flags.append("POSITIVE_TONE")
        details.append(f"Pozitif ifadeler ({positive_count}): +{bonus}")
    
    # ============================================================
    # 5. BİLGİLENDİRİCİ İÇERİK (REGEX)
    # ============================================================
    info_count = 0
    for pattern in keywords["informative_patterns"]:
        if re.search(pattern, response_lower):
            info_count += 1
    
    if info_count > 0:
        bonus = min(info_count * 5, 15)  # Max +15
        score += bonus
        flags.append("INFORMATIVE")
        details.append(f"Bilgilendirici içerik ({info_count}): +{bonus}")
    
    # ============================================================
    # 6. PROFESYONELLİK
    # ============================================================
    prof_count = sum(1 for p in keywords["professional"] if p in response_lower)
    if prof_count > 0:
        score += 5
        flags.append("PROFESSIONAL")
        details.append("Profesyonel ifade: +5")
    
    # ============================================================
    # 7. YÖNLENDİRME (Nötr - ceza veya bonus yok, sadece flag)
    # ============================================================
    has_redirect = any(r in response_lower for r in keywords["redirect"])
    if has_redirect:
        flags.append("REDIRECT")
        details.append("Yönlendirme içeriyor: 0")
    
    # ============================================================
    # 8. KİBAR RED (Nötr - "maalesef" ceza DEĞİL!)
    # ============================================================
    has_polite_decline = any(p in response_lower for p in keywords["polite_decline"])
    if has_polite_decline:
        flags.append("POLITE_DECLINE")
        details.append("Kibar red: 0 (normal)")
        # NOT: Artık ceza yok!
    
    # ============================================================
    # 9. HATA KONTROLÜ (Gerçek sistem hataları)
    # ============================================================
    has_error = any(e in response_lower for e in keywords["error"])
    if has_error:
        score -= 20
        flags.append("ERROR_DETECTED")
        details.append("Hata mesajı: -20")
    
    # ============================================================
    # 10. DİL TUTARLILIĞI
    # ============================================================
    turkish_chars = set("çğıöşüÇĞİÖŞÜ")
    has_turkish = any(c in response for c in turkish_chars)
    
    if lang == "tr":
        if has_turkish:
            score += 5
            flags.append("LANG_TR_OK")
            details.append("Türkçe tutarlı: +5")
        elif response_len > 50:
            # Türkçe karakter yok ama uzun yanıt - muhtemelen sistem mesajı
            flags.append("LANG_CHECK")
    elif lang == "en":
        if not has_turkish and response_len > 20:
            score += 5
            flags.append("LANG_EN_OK")
            details.append("İngilizce tutarlı: +5")
    
    # ============================================================
    # 11. TAKİP SORUSU
    # ============================================================
    if "?" in response:
        # Yanıtta soru var - etkileşimli
        score += 3
        flags.append("INTERACTIVE")
        details.append("Takip sorusu: +3")
    
    # ============================================================
    # 12. BAĞLAMSAL UYUM (Soru-Cevap ilişkisi)
    # ============================================================
    # Soruda geçen anahtar kelimelerin yanıtta da olup olmadığını kontrol et
    question_keywords = set(re.findall(r'\b\w{4,}\b', question_lower))
    response_keywords = set(re.findall(r'\b\w{4,}\b', response_lower))
    
    common_keywords = question_keywords & response_keywords
    if len(common_keywords) >= 2:
        score += 5
        flags.append("CONTEXTUAL")
        details.append("Bağlamsal uyum: +5")
    
    # ============================================================
    # FINAL SKOR
    # ============================================================
    score = max(0, min(100, score))
    
    return {
        "score": score,
        "flags": flags,
        "details": details,
        "breakdown": {
            "base": 50,
            "final": score,
            "adjustments": details
        }
    }


async def load_questions() -> Dict:
    """Test sorularını yükle"""
    with open(QUESTIONS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_question_text(question_item, lang: str = "tr") -> str:
    """
    Soru metnini çıkar - hem eski hem yeni format destekler.
    """
    if isinstance(question_item, str):
        return question_item
    elif isinstance(question_item, dict):
        return question_item.get(lang, question_item.get("tr", ""))
    else:
        return str(question_item)


async def run_single_test(client: httpx.AsyncClient, phone: str, question: str, category: str, lang: str = "tr") -> Dict:
    """Tek bir test sorusunu çalıştır"""
    start_time = time.time()
    
    try:
        response = await client.post(
            f"{BASE_URL}{TEST_ENDPOINT}",
            params={"phone": phone, "message": question},
            timeout=30.0
        )
        elapsed = time.time() - start_time
        
        if response.status_code == 200:
            data = response.json()
            bot_response = data.get("response", "")
            response_source = data.get("response_source", "UNKNOWN")
            
            # GELİŞTİRİLMİŞ SKORLAMA
            quality = evaluate_response_v21(question, bot_response, lang)
            
            return {
                "status": "success",
                "question": question,
                "language": lang,
                "category": category,
                "response": bot_response,
                "response_source": response_source,
                "response_time": round(elapsed, 3),
                "response_length": len(bot_response),
                "quality_score": quality["score"],
                "quality_flags": quality["flags"],
                "score_details": quality["details"],
                "timestamp": datetime.now().isoformat()
            }
        else:
            return {
                "status": "http_error",
                "question": question,
                "language": lang,
                "category": category,
                "response": f"HTTP {response.status_code}",
                "response_source": "ERROR",
                "response_time": round(elapsed, 3),
                "response_length": 0,
                "quality_score": 0,
                "quality_flags": ["HTTP_ERROR"],
                "score_details": ["HTTP hatası: 0"],
                "timestamp": datetime.now().isoformat()
            }
    except Exception as e:
        elapsed = time.time() - start_time
        return {
            "status": "exception",
            "question": question,
            "language": lang,
            "category": category,
            "response": str(e),
            "response_source": "ERROR",
            "response_time": round(elapsed, 3),
            "response_length": 0,
            "quality_score": 0,
            "quality_flags": ["EXCEPTION"],
            "score_details": [f"Exception: {str(e)}"],
            "timestamp": datetime.now().isoformat()
        }


async def run_all_tests(test_langs: List[str] = ["tr"], progress_callback=None, max_per_category: int = None) -> Dict:
    """Tüm testleri çalıştır
    
    Args:
        test_langs: Test edilecek diller ["tr"], ["en"], veya ["tr", "en"]
        progress_callback: İlerleme callback fonksiyonu
        max_per_category: Her kategoriden max kaç soru (None = hepsi)
    """
    questions_data = await load_questions()
    results = {
        "metadata": {
            "test_date": datetime.now().isoformat(),
            "base_url": BASE_URL,
            "test_languages": test_langs,
            "scoring_version": "2.1",
            "total_categories": len(questions_data["categories"]),
            "total_questions": 0,
            "total_success": 0,
            "total_errors": 0,
            "avg_response_time": 0,
            "avg_quality_score": 0,
            "score_distribution": {
                "excellent": 0,   # 90-100
                "good": 0,        # 75-89
                "average": 0,     # 60-74
                "poor": 0,        # 40-59
                "critical": 0     # 0-39
            }
        },
        "categories": {},
        "all_results": []
    }
    
    total_time = 0
    total_score = 0
    question_count = 0
    
    async with httpx.AsyncClient() as client:
        for cat_key, cat_data in questions_data["categories"].items():
            cat_results = []
            cat_name = cat_data.get("name_tr", cat_key)
            questions = cat_data["questions"]
            
            # Hızlı test modu: sadece belirli sayıda soru al
            if max_per_category:
                questions = questions[:max_per_category]
            
            for i, question_item in enumerate(questions):
                for lang in test_langs:
                    question_text = extract_question_text(question_item, lang)
                    
                    if not question_text:
                        continue
                    
                    question_count += 1
                    phone = f"TEST_{cat_key.upper()}_{i:03d}_{lang.upper()}"
                    
                    if progress_callback:
                        progress_callback(question_count, cat_name, question_text[:50], lang)
                    
                    result = await run_single_test(client, phone, question_text, cat_key, lang)
                    cat_results.append(result)
                    results["all_results"].append(result)
                    
                    if result["status"] == "success":
                        results["metadata"]["total_success"] += 1
                    else:
                        results["metadata"]["total_errors"] += 1
                    
                    total_time += result["response_time"]
                    total_score += result["quality_score"]
                    
                    # Skor dağılımı
                    score = result["quality_score"]
                    if score >= 90:
                        results["metadata"]["score_distribution"]["excellent"] += 1
                    elif score >= 75:
                        results["metadata"]["score_distribution"]["good"] += 1
                    elif score >= 60:
                        results["metadata"]["score_distribution"]["average"] += 1
                    elif score >= 40:
                        results["metadata"]["score_distribution"]["poor"] += 1
                    else:
                        results["metadata"]["score_distribution"]["critical"] += 1
                    
                    await asyncio.sleep(0.1)
            
            if cat_results:
                results["categories"][cat_key] = {
                    "name": cat_name,
                    "total": len(cat_results),
                    "success": sum(1 for r in cat_results if r["status"] == "success"),
                    "avg_time": round(sum(r["response_time"] for r in cat_results) / len(cat_results), 3),
                    "avg_score": round(sum(r["quality_score"] for r in cat_results) / len(cat_results), 1),
                    "results": cat_results
                }
    
    results["metadata"]["total_questions"] = question_count
    results["metadata"]["avg_response_time"] = round(total_time / question_count, 3) if question_count > 0 else 0
    results["metadata"]["avg_quality_score"] = round(total_score / question_count, 1) if question_count > 0 else 0
    
    return results


def generate_json_report(results: Dict, output_path: Path):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON raporu: {output_path}")


def generate_markdown_report(results: Dict, output_path: Path):
    """Markdown rapor üret - TAM YANITLAR + SKOR DETAYLARI"""
    md = []
    md.append("# 📊 KassandraOpenAI Sistem Test Raporu v2.1\n")
    md.append(f"**Test Tarihi:** {results['metadata']['test_date'][:19]}\n")
    md.append(f"**Backend URL:** {results['metadata']['base_url']}\n")
    md.append(f"**Test Dilleri:** {', '.join(results['metadata']['test_languages'])}\n")
    md.append(f"**Skorlama Versiyonu:** {results['metadata']['scoring_version']}\n")
    
    # Genel Özet
    md.append("\n## 📈 Genel Özet\n")
    md.append("| Metrik | Değer |")
    md.append("|--------|-------|")
    md.append(f"| Toplam Soru | {results['metadata']['total_questions']} |")
    md.append(f"| Başarılı | {results['metadata']['total_success']} |")
    md.append(f"| Hatalı | {results['metadata']['total_errors']} |")
    md.append(f"| Başarı Oranı | {round(results['metadata']['total_success']/results['metadata']['total_questions']*100, 1)}% |")
    md.append(f"| Ort. Yanıt Süresi | {results['metadata']['avg_response_time']}s |")
    md.append(f"| Ort. Kalite Skoru | {results['metadata']['avg_quality_score']}/100 |")
    
    # Skor Dağılımı
    md.append("\n## 📊 Skor Dağılımı\n")
    dist = results['metadata']['score_distribution']
    md.append("| Seviye | Aralık | Sayı | Oran |")
    md.append("|--------|--------|------|------|")
    total = results['metadata']['total_questions']
    md.append(f"| 🌟 Mükemmel | 90-100 | {dist['excellent']} | {round(dist['excellent']/total*100, 1)}% |")
    md.append(f"| ✅ İyi | 75-89 | {dist['good']} | {round(dist['good']/total*100, 1)}% |")
    md.append(f"| 📊 Ortalama | 60-74 | {dist['average']} | {round(dist['average']/total*100, 1)}% |")
    md.append(f"| ⚠️ Zayıf | 40-59 | {dist['poor']} | {round(dist['poor']/total*100, 1)}% |")
    md.append(f"| ❌ Kritik | 0-39 | {dist['critical']} | {round(dist['critical']/total*100, 1)}% |")
    
    # Kategori Özeti
    md.append("\n## 📁 Kategori Bazlı Sonuçlar\n")
    md.append("| Kategori | Toplam | Başarılı | Ort. Süre | Ort. Skor |")
    md.append("|----------|--------|----------|-----------|-----------|")
    
    for cat_key, cat_data in results["categories"].items():
        score_emoji = "🌟" if cat_data['avg_score'] >= 90 else "✅" if cat_data['avg_score'] >= 75 else "📊" if cat_data['avg_score'] >= 60 else "⚠️"
        md.append(f"| {cat_data['name']} | {cat_data['total']} | {cat_data['success']} | {cat_data['avg_time']}s | {score_emoji} {cat_data['avg_score']} |")
    
    # Skorlama Kriterleri Açıklaması
    md.append("\n## 📐 Skorlama Kriterleri (v2.1)\n")
    md.append("| Kriter | Puan | Açıklama |")
    md.append("|--------|------|----------|")
    md.append("| Başlangıç | 50 | Temel skor |")
    md.append("| Uygun uzunluk (50-500 kar.) | +15 | İdeal yanıt uzunluğu |")
    md.append("| Selamlama | +5 | Merhaba, Hoş geldiniz |")
    md.append("| Profesyonel kapanış | +5 | İyi günler, Yardımcı olabilirim |")
    md.append("| Pozitif ifadeler | +3-10 | Evet, mümkün, yapabiliriz |")
    md.append("| Bilgilendirici içerik | +5-15 | Tarih, saat, fiyat bilgisi |")
    md.append("| Profesyonel ton | +5 | Yardımcı olabilirim |")
    md.append("| Dil tutarlılığı | +5 | TR/EN uyumu |")
    md.append("| Takip sorusu | +3 | Etkileşimli yanıt |")
    md.append("| Bağlamsal uyum | +5 | Soru-cevap ilişkisi |")
    md.append("| Kibar red | 0 | Maalesef, üzgünüm (CEZA YOK) |")
    md.append("| Sistem hatası | -20 | Hata mesajı |")
    md.append("| Çok kısa (<20 kar.) | -25 | Yetersiz yanıt |")
    
    # Detaylı Sonuçlar
    md.append("\n## 📝 Tüm Detaylı Sonuçlar\n")
    
    for cat_key, cat_data in results["categories"].items():
        md.append(f"\n### {cat_data['name']}\n")
        md.append(f"*({cat_data['total']} soru, ort. skor: {cat_data['avg_score']})*\n")
        
        for i, r in enumerate(cat_data["results"]):
            status_emoji = "✅" if r["status"] == "success" else "❌"
            lang_badge = f"[{r.get('language', 'TR').upper()}]"
            score = r['quality_score']
            score_emoji = "🌟" if score >= 90 else "✅" if score >= 75 else "📊" if score >= 60 else "⚠️" if score >= 40 else "❌"
            
            md.append(f"\n---\n")
            md.append(f"#### {i+1}. {status_emoji} {lang_badge} Soru\n")
            md.append(f"**Soru:** {r['question']}\n")
            md.append(f"\n**Yanıt:**\n")
            md.append(f"```\n{r['response']}\n```\n")
            
            # Eğer skor düşükse veya hata varsa DOĞRU YANIT bölümü ekle
            if score < 75 or r["status"] != "success":
                md.append(f"```\n")
                md.append(f"YANITIN YANLIŞ OLMA SEBEBİ : \"\"\n")
                md.append(f"```\n")
                md.append(f"```\n")
                md.append(f"DOĞRU YANIT : \"\"\n")
                md.append(f"```\n")
            
            md.append(f"\n| Kaynak | Süre | Skor | Flags |")
            md.append(f"|--------|------|------|-------|")
            md.append(f"| {r['response_source']} | {r['response_time']}s | {score_emoji} {score}/100 | {', '.join(r['quality_flags']) if r['quality_flags'] else '-'} |")
            
            # Skor detayları
            if r.get('score_details'):
                md.append(f"\n<details><summary>📐 Skor Detayı</summary>\n")
                for detail in r['score_details']:
                    md.append(f"- {detail}")
                md.append(f"</details>\n")
    
    # Sorunlu Yanıtlar
    md.append("\n## ⚠️ Dikkat Gerektiren Yanıtlar (Skor < 75)\n")
    problem_results = [r for r in results["all_results"] if r["quality_score"] < 75 or r["status"] != "success"]
    
    if problem_results:
        md.append("| # | Dil | Kategori | Soru | Skor | Flags |")
        md.append("|---|-----|----------|------|------|-------|")
        for i, r in enumerate(problem_results[:50], 1):  # İlk 50
            flags = ", ".join(r["quality_flags"][:3]) if r["quality_flags"] else r["status"]
            lang = r.get("language", "tr").upper()
            score_emoji = "⚠️" if r['quality_score'] >= 40 else "❌"
            md.append(f"| {i} | {lang} | {r['category']} | {r['question'][:35]}... | {score_emoji} {r['quality_score']} | {flags} |")
        
        if len(problem_results) > 50:
            md.append(f"\n*...ve {len(problem_results) - 50} tane daha*")
    else:
        md.append("✅ Tüm yanıtlar 75 ve üzeri skor aldı!")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md))
    print(f"✅ Markdown raporu: {output_path}")


def generate_excel_report(results: Dict, output_path: Path):
    """Excel rapor üret - TAM YANITLAR + SKOR DETAYLARI"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️ openpyxl yüklü değil. Yüklemek için: pip install openpyxl")
        return
    
    wb = Workbook()
    
    # Renkler
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    excellent_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Yeşil
    good_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")  # Açık yeşil
    average_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Sarı
    poor_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Pembe
    critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Kırmızı
    
    # === Sheet 1: Özet ===
    ws_summary = wb.active
    ws_summary.title = "Özet"
    
    summary_data = [
        ["Metrik", "Değer"],
        ["Test Tarihi", results['metadata']['test_date'][:19]],
        ["Skorlama Versiyonu", results['metadata']['scoring_version']],
        ["Test Dilleri", ", ".join(results['metadata']['test_languages'])],
        ["Toplam Soru", results['metadata']['total_questions']],
        ["Başarılı", results['metadata']['total_success']],
        ["Hatalı", results['metadata']['total_errors']],
        ["Başarı Oranı (%)", round(results['metadata']['total_success']/results['metadata']['total_questions']*100, 1)],
        ["Ort. Yanıt Süresi (s)", results['metadata']['avg_response_time']],
        ["Ort. Kalite Skoru", results['metadata']['avg_quality_score']],
        ["", ""],
        ["Skor Dağılımı", ""],
        ["Mükemmel (90-100)", results['metadata']['score_distribution']['excellent']],
        ["İyi (75-89)", results['metadata']['score_distribution']['good']],
        ["Ortalama (60-74)", results['metadata']['score_distribution']['average']],
        ["Zayıf (40-59)", results['metadata']['score_distribution']['poor']],
        ["Kritik (0-39)", results['metadata']['score_distribution']['critical']],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 25
    
    # === Sheet 2: Kategoriler ===
    ws_cat = wb.create_sheet("Kategoriler")
    cat_headers = ["Kategori", "Toplam", "Başarılı", "Hatalı", "Başarı %", "Ort. Süre (s)", "Ort. Skor"]
    
    for col_idx, header in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row_idx = 2
    for cat_key, cat_data in results["categories"].items():
        ws_cat.cell(row=row_idx, column=1, value=cat_data['name'])
        ws_cat.cell(row=row_idx, column=2, value=cat_data['total'])
        ws_cat.cell(row=row_idx, column=3, value=cat_data['success'])
        ws_cat.cell(row=row_idx, column=4, value=cat_data['total'] - cat_data['success'])
        ws_cat.cell(row=row_idx, column=5, value=round(cat_data['success']/cat_data['total']*100, 1))
        ws_cat.cell(row=row_idx, column=6, value=cat_data['avg_time'])
        score_cell = ws_cat.cell(row=row_idx, column=7, value=cat_data['avg_score'])
        
        # Skor renklendirme
        if cat_data['avg_score'] >= 90:
            score_cell.fill = excellent_fill
        elif cat_data['avg_score'] >= 75:
            score_cell.fill = good_fill
        elif cat_data['avg_score'] >= 60:
            score_cell.fill = average_fill
        elif cat_data['avg_score'] >= 40:
            score_cell.fill = poor_fill
        else:
            score_cell.fill = critical_fill
        
        row_idx += 1
    
    for col_idx in range(1, 8):
        ws_cat.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # === Sheet 3: Tüm Sonuçlar ===
    ws_all = wb.create_sheet("Tüm Sonuçlar")
    all_headers = ["#", "Dil", "Kategori", "Soru", "Yanıt (TAM)", "Kaynak", "Süre (s)", "Skor", "Durum", "Flags", "Skor Detayı"]
    
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_idx, r in enumerate(results["all_results"], 2):
        ws_all.cell(row=row_idx, column=1, value=row_idx-1)
        ws_all.cell(row=row_idx, column=2, value=r.get('language', 'tr').upper())
        ws_all.cell(row=row_idx, column=3, value=r['category'])
        ws_all.cell(row=row_idx, column=4, value=r['question'])
        ws_all.cell(row=row_idx, column=5, value=r['response'])
        ws_all.cell(row=row_idx, column=6, value=r['response_source'])
        ws_all.cell(row=row_idx, column=7, value=r['response_time'])
        score_cell = ws_all.cell(row=row_idx, column=8, value=r['quality_score'])
        ws_all.cell(row=row_idx, column=9, value=r['status'])
        ws_all.cell(row=row_idx, column=10, value=", ".join(r['quality_flags']))
        ws_all.cell(row=row_idx, column=11, value=" | ".join(r.get('score_details', [])))
        
        # Hücre hizalama
        ws_all.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True, vertical='top')
        ws_all.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical='top')
        ws_all.cell(row=row_idx, column=11).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Skor renklendirme
        score = r['quality_score']
        if score >= 90:
            score_cell.fill = excellent_fill
        elif score >= 75:
            score_cell.fill = good_fill
        elif score >= 60:
            score_cell.fill = average_fill
        elif score >= 40:
            score_cell.fill = poor_fill
        else:
            score_cell.fill = critical_fill
    
    # Sütun genişlikleri
    ws_all.column_dimensions['A'].width = 5
    ws_all.column_dimensions['B'].width = 6
    ws_all.column_dimensions['C'].width = 20
    ws_all.column_dimensions['D'].width = 50
    ws_all.column_dimensions['E'].width = 80
    ws_all.column_dimensions['F'].width = 10
    ws_all.column_dimensions['G'].width = 10
    ws_all.column_dimensions['H'].width = 8
    ws_all.column_dimensions['I'].width = 10
    ws_all.column_dimensions['J'].width = 30
    ws_all.column_dimensions['K'].width = 50
    
    for row_idx in range(2, len(results["all_results"]) + 2):
        ws_all.row_dimensions[row_idx].height = 60
    
    wb.save(output_path)
    print(f"✅ Excel raporu: {output_path}")


async def main():
    # Komut satırı argümanları
    test_langs = ["tr"]
    quick_mode = "--quick" in sys.argv
    sample_count = 1 if quick_mode else None  # Her kategoriden kaç soru
    
    if "--both" in sys.argv:
        test_langs = ["tr", "en"]
        print("📌 Mod: Hem Türkçe hem İngilizce sorular test edilecek")
    elif "--en" in sys.argv:
        test_langs = ["en"]
        print("📌 Mod: Sadece İngilizce sorular test edilecek")
    else:
        print("📌 Mod: Sadece Türkçe sorular test edilecek")
        print("   (--both veya --en parametresi ile değiştirebilirsiniz)")
    
    if quick_mode:
        print("⚡ HIZLI TEST: Her kategoriden sadece 1 soru")
    
    print("=" * 60)
    print("🧪 KassandraOpenAI Sistem Testi v2.1 Başlıyor")
    print("   GELİŞTİRİLMİŞ SKORLAMA SİSTEMİ")
    print("=" * 60)
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    def progress(count, category, question, lang):
        lang_badge = f"[{lang.upper()}]"
        print(f"[{count:3d}] {lang_badge} {category}: {question}...")
    
    print("\n📋 Testler çalıştırılıyor...\n")
    start_time = time.time()
    
    results = await run_all_tests(test_langs=test_langs, progress_callback=progress, max_per_category=sample_count)
    
    total_time = time.time() - start_time
    print(f"\n⏱️ Toplam test süresi: {round(total_time, 1)} saniye")
    
    print("\n📊 Raporlar oluşturuluyor...\n")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    generate_json_report(results, OUTPUT_DIR / f"test_results_{timestamp}.json")
    generate_markdown_report(results, OUTPUT_DIR / f"test_report_{timestamp}.md")
    generate_excel_report(results, OUTPUT_DIR / f"test_report_{timestamp}.xlsx")
    
    # Özet
    print("\n" + "=" * 60)
    print("📈 TEST ÖZETİ (Skorlama v2.1)")
    print("=" * 60)
    print(f"Test Dilleri: {', '.join(test_langs)}")
    print(f"Toplam Soru: {results['metadata']['total_questions']}")
    print(f"Başarılı: {results['metadata']['total_success']}")
    print(f"Hatalı: {results['metadata']['total_errors']}")
    print(f"Başarı Oranı: {round(results['metadata']['total_success']/results['metadata']['total_questions']*100, 1)}%")
    print(f"Ort. Yanıt Süresi: {results['metadata']['avg_response_time']}s")
    print(f"Ort. Kalite Skoru: {results['metadata']['avg_quality_score']}/100")
    print("-" * 60)
    print("📊 SKOR DAĞILIMI:")
    dist = results['metadata']['score_distribution']
    print(f"   🌟 Mükemmel (90-100): {dist['excellent']}")
    print(f"   ✅ İyi (75-89):       {dist['good']}")
    print(f"   📊 Ortalama (60-74):  {dist['average']}")
    print(f"   ⚠️ Zayıf (40-59):     {dist['poor']}")
    print(f"   ❌ Kritik (0-39):     {dist['critical']}")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
